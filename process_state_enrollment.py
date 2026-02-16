"""
Process state enrollment projections from multiple sources.
- NCES Digest Table 203.20: state-level projections for all 50 states + DC (2022-2031)
- California DOF: county-level projections (2024-25 to 2046-47)
- Additional state sources as available

Outputs:
- state_enrollment_projections.json: combined dataset
- data_coverage_report.md: summary of coverage and gaps
"""

import json
import os
import openpyxl
import math

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
OUT_DIR = os.path.dirname(__file__)

# State FIPS codes and abbreviations
STATE_INFO = {
    'Alabama': ('01', 'AL'), 'Alaska': ('02', 'AK'), 'Arizona': ('04', 'AZ'),
    'Arkansas': ('05', 'AR'), 'California': ('06', 'CA'), 'Colorado': ('08', 'CO'),
    'Connecticut': ('09', 'CT'), 'Delaware': ('10', 'DE'), 'District of Columbia': ('11', 'DC'),
    'Florida': ('12', 'FL'), 'Georgia': ('13', 'GA'), 'Hawaii': ('15', 'HI'),
    'Idaho': ('16', 'ID'), 'Illinois': ('17', 'IL'), 'Indiana': ('18', 'IN'),
    'Iowa': ('19', 'IA'), 'Kansas': ('20', 'KS'), 'Kentucky': ('21', 'KY'),
    'Louisiana': ('22', 'LA'), 'Maine': ('23', 'ME'), 'Maryland': ('24', 'MD'),
    'Massachusetts': ('25', 'MA'), 'Michigan': ('26', 'MI'), 'Minnesota': ('27', 'MN'),
    'Mississippi': ('28', 'MS'), 'Missouri': ('29', 'MO'), 'Montana': ('30', 'MT'),
    'Nebraska': ('31', 'NE'), 'Nevada': ('32', 'NV'), 'New Hampshire': ('33', 'NH'),
    'New Jersey': ('34', 'NJ'), 'New Mexico': ('35', 'NM'), 'New York': ('36', 'NY'),
    'North Carolina': ('37', 'NC'), 'North Dakota': ('38', 'ND'), 'Ohio': ('39', 'OH'),
    'Oklahoma': ('40', 'OK'), 'Oregon': ('41', 'OR'), 'Pennsylvania': ('42', 'PA'),
    'Rhode Island': ('44', 'RI'), 'South Carolina': ('45', 'SC'), 'South Dakota': ('46', 'SD'),
    'Tennessee': ('47', 'TN'), 'Texas': ('48', 'TX'), 'Utah': ('49', 'UT'),
    'Vermont': ('50', 'VT'), 'Virginia': ('51', 'VA'), 'Washington': ('53', 'WA'),
    'West Virginia': ('54', 'WV'), 'Wisconsin': ('55', 'WI'), 'Wyoming': ('56', 'WY'),
}

# California county FIPS codes
CA_COUNTY_FIPS = {
    'Alameda': '06001', 'Alpine': '06003', 'Amador': '06005', 'Butte': '06007',
    'Calaveras': '06009', 'Colusa': '06011', 'Contra Costa': '06013', 'Del Norte': '06015',
    'El Dorado': '06017', 'Fresno': '06019', 'Glenn': '06021', 'Humboldt': '06023',
    'Imperial': '06025', 'Inyo': '06027', 'Kern': '06029', 'Kings': '06031',
    'Lake': '06033', 'Lassen': '06035', 'Los Angeles': '06037', 'Madera': '06039',
    'Marin': '06041', 'Mariposa': '06043', 'Mendocino': '06045', 'Merced': '06047',
    'Modoc': '06049', 'Mono': '06051', 'Monterey': '06053', 'Napa': '06055',
    'Nevada': '06057', 'Orange': '06059', 'Placer': '06061', 'Plumas': '06063',
    'Riverside': '06065', 'Sacramento': '06067', 'San Benito': '06069',
    'San Bernardino': '06071', 'San Diego': '06073', 'San Francisco': '06075',
    'San Joaquin': '06077', 'San Luis Obispo': '06079', 'San Mateo': '06081',
    'Santa Barbara': '06083', 'Santa Clara': '06085', 'Santa Cruz': '06087',
    'Shasta': '06089', 'Sierra': '06091', 'Siskiyou': '06093', 'Solano': '06095',
    'Sonoma': '06097', 'Stanislaus': '06099', 'Sutter': '06101', 'Tehama': '06103',
    'Trinity': '06105', 'Tulare': '06107', 'Tuolumne': '06109', 'Ventura': '06111',
    'Yolo': '06113', 'Yuba': '06115',
}


def process_nces():
    """Process NCES state-level enrollment projections."""
    print("Processing NCES state-level projections...")

    # Load from state_sources.json which has the extracted data
    with open(os.path.join(OUT_DIR, 'state_sources.json'), 'r') as f:
        sources = json.load(f)

    nces_data = sources['national_baseline']['projected_enrollment']

    records = []
    for state_name, (fips, abbr) in STATE_INFO.items():
        if abbr not in nces_data:
            continue
        proj = nces_data[abbr]

        # NCES provides: 2022, 2023, 2024, 2025, 2031
        # Interpolate 2026-2030 linearly between 2025 and 2031
        years = {}
        for y_str, val in proj.items():
            years[int(y_str)] = val

        # Fill in 2026-2030 via linear interpolation
        if 2025 in years and 2031 in years:
            v2025 = years[2025]
            v2031 = years[2031]
            for y in range(2026, 2031):
                frac = (y - 2025) / (2031 - 2025)
                years[y] = round(v2025 + frac * (v2031 - v2025))

        # Also extend to 2050 using 2025-2031 trend
        if 2025 in years and 2031 in years:
            annual_change = (years[2031] - years[2025]) / 6
            for y in range(2032, 2051):
                years[y] = max(0, round(years[2031] + annual_change * (y - 2031)))

        for year, enrollment in sorted(years.items()):
            records.append({
                'state_fips': fips,
                'state_name': state_name,
                'state_abbr': abbr,
                'county_fips': None,
                'geo_name': state_name,
                'geo_level': 'state',
                'year': year,
                'projected_enrollment': enrollment,
                'source': 'NCES Projections of Education Statistics to 2031',
                'interpolated': year in range(2026, 2031),
                'extrapolated': year > 2031,
            })

    print(f"  NCES: {len(records)} records for {len(STATE_INFO)} states")
    return records


def process_california():
    """Process California DOF county-level enrollment projections."""
    print("Processing California DOF county-level projections...")

    xlsx_path = os.path.join(DATA_DIR, 'ca_k12_enrollment_2025.xlsx')
    if not os.path.exists(xlsx_path):
        print("  CA DOF file not found, skipping")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['County Enrollment Projection']

    # Read header row (row 3) for year labels
    year_cols = {}
    for c in range(2, 25):
        val = ws.cell(row=3, column=c).value
        if val and isinstance(val, str) and '-' in val:
            # e.g. '2024-25' -> use first year as the fall year
            fall_year = int(val.split('-')[0])
            year_cols[c] = fall_year

    records = []
    for r in range(4, 62):
        county_name = ws.cell(row=r, column=1).value
        if not county_name or county_name == 'California':
            continue

        county_fips = CA_COUNTY_FIPS.get(county_name)
        if not county_fips:
            print(f"  Warning: no FIPS for CA county '{county_name}'")
            continue

        for col, year in year_cols.items():
            val = ws.cell(row=r, column=col).value
            if val is not None:
                records.append({
                    'state_fips': '06',
                    'state_name': 'California',
                    'state_abbr': 'CA',
                    'county_fips': county_fips,
                    'geo_name': f"{county_name} County, CA",
                    'geo_level': 'county',
                    'year': year,
                    'projected_enrollment': int(val),
                    'source': 'California DOF K-12 Enrollment Projections, 2025 Series',
                    'interpolated': False,
                    'extrapolated': False,
                })

    print(f"  California: {len(records)} records for {len(CA_COUNTY_FIPS)} counties")
    return records


# Iowa county FIPS lookup (county code in IA data is 2-digit, maps to 5-digit FIPS)
IA_COUNTY_FIPS = {
    '01': '19001', '02': '19003', '03': '19005', '04': '19007', '05': '19009',
    '06': '19011', '07': '19013', '08': '19015', '09': '19017', '10': '19019',
    '11': '19021', '12': '19023', '13': '19025', '14': '19027', '15': '19029',
    '16': '19031', '17': '19033', '18': '19035', '19': '19037', '20': '19039',
    '21': '19041', '22': '19043', '23': '19045', '24': '19047', '25': '19049',
    '26': '19051', '27': '19053', '28': '19055', '29': '19057', '30': '19059',
    '31': '19061', '32': '19063', '33': '19065', '34': '19067', '35': '19069',
    '36': '19071', '37': '19073', '38': '19075', '39': '19077', '40': '19079',
    '41': '19081', '42': '19083', '43': '19085', '44': '19087', '45': '19089',
    '46': '19091', '47': '19093', '48': '19095', '49': '19097', '50': '19099',
    '51': '19101', '52': '19103', '53': '19105', '54': '19107', '55': '19109',
    '56': '19111', '57': '19113', '58': '19115', '59': '19117', '60': '19119',
    '61': '19121', '62': '19123', '63': '19125', '64': '19127', '65': '19129',
    '66': '19131', '67': '19133', '68': '19135', '69': '19137', '70': '19139',
    '71': '19141', '72': '19143', '73': '19145', '74': '19147', '75': '19149',
    '76': '19151', '77': '19153', '78': '19155', '79': '19157', '80': '19159',
    '81': '19161', '82': '19163', '83': '19165', '84': '19167', '85': '19169',
    '86': '19171', '87': '19173', '88': '19175', '89': '19177', '90': '19179',
    '91': '19181', '92': '19183', '93': '19185', '94': '19187', '95': '19189',
    '96': '19191', '97': '19193', '98': '19195', '99': '19197',
}

# Maryland county FIPS
MD_COUNTY_FIPS = {
    'Anne Arundel County': '24003', 'Baltimore County': '24005',
    'Carroll County': '24013', 'Harford County': '24025',
    'Howard County': '24027', 'Baltimore City': '24510',
    'Frederick County': '24021', 'Montgomery County': '24031',
    "Prince George's County": '24033', 'Calvert County': '24009',
    'Charles County': '24017', "St. Mary's County": '24037',
    'Allegany County': '24001', 'Garrett County': '24023',
    'Washington County': '24043', 'Caroline County': '24011',
    'Cecil County': '24015', 'Kent County': '24029',
    "Queen Anne's County": '24035', 'Talbot County': '24041',
    'Dorchester County': '24019', 'Somerset County': '24039',
    'Wicomico County': '24045', 'Worcester County': '24047',
}

# Pennsylvania county name -> FIPS
PA_COUNTY_FIPS = {
    'Adams': '42001', 'Allegheny': '42003', 'Armstrong': '42005', 'Beaver': '42007',
    'Bedford': '42009', 'Berks': '42011', 'Blair': '42013', 'Bradford': '42015',
    'Bucks': '42017', 'Butler': '42019', 'Cambria': '42021', 'Cameron': '42023',
    'Carbon': '42025', 'Centre': '42027', 'Chester': '42029', 'Clarion': '42031',
    'Clearfield': '42033', 'Clinton': '42035', 'Columbia': '42037', 'Crawford': '42039',
    'Cumberland': '42041', 'Dauphin': '42043', 'Delaware': '42045', 'Elk': '42047',
    'Erie': '42049', 'Fayette': '42051', 'Forest': '42053', 'Franklin': '42055',
    'Fulton': '42057', 'Greene': '42059', 'Huntingdon': '42061', 'Indiana': '42063',
    'Jefferson': '42065', 'Juniata': '42067', 'Lackawanna': '42069', 'Lancaster': '42071',
    'Lawrence': '42073', 'Lebanon': '42075', 'Lehigh': '42077', 'Luzerne': '42079',
    'Lycoming': '42081', 'McKean': '42083', 'Mercer': '42085', 'Mifflin': '42087',
    'Monroe': '42089', 'Montgomery': '42091', 'Montour': '42093', 'Northampton': '42095',
    'Northumberland': '42097', 'Perry': '42099', 'Philadelphia': '42101', 'Pike': '42103',
    'Potter': '42105', 'Schuylkill': '42107', 'Snyder': '42109', 'Somerset': '42111',
    'Sullivan': '42113', 'Susquehanna': '42115', 'Tioga': '42117', 'Union': '42119',
    'Venango': '42121', 'Warren': '42123', 'Washington': '42125', 'Wayne': '42127',
    'Westmoreland': '42129', 'Wyoming': '42131', 'York': '42133',
}


def process_iowa():
    """Process Iowa district-level enrollment projections, aggregated to county."""
    print("Processing Iowa enrollment projections...")

    xlsx_path = os.path.join(DATA_DIR, 'ia_enrollment_projections.xlsx')
    if not os.path.exists(xlsx_path):
        print("  Iowa file not found, skipping")
        return []

    import warnings
    warnings.filterwarnings('ignore')

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Headers at row 9: AEA Code, AEA Name, County Code, County Name, District Code, District Name,
    # then years: 2020-21 through 2024-25 (actual), 2025-26 through 2029-30 (projected)
    year_labels = []
    for c in range(7, 17):
        val = ws.cell(row=9, column=c).value
        year_labels.append(val)

    # Aggregate district enrollment to county level
    # county_data[county_code] = {county_name, year_totals: {year: total}}
    county_data = {}
    for r in range(10, ws.max_row + 1):
        county_code = ws.cell(row=r, column=3).value
        county_name = ws.cell(row=r, column=4).value
        if not county_code or not county_name:
            continue

        county_code = str(county_code).zfill(2)
        if county_code not in county_data:
            county_data[county_code] = {'name': county_name, 'years': {}}

        for i, c in enumerate(range(7, 17)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                year_label = year_labels[i]
                if year_label not in county_data[county_code]['years']:
                    county_data[county_code]['years'][year_label] = 0
                county_data[county_code]['years'][year_label] += val

    # Build records (projected years only: 2025-26 through 2029-30)
    projected_years = {'2025-26', '2026-27', '2027-28', '2028-29', '2029-30'}
    records = []
    for county_code, info in county_data.items():
        county_fips = IA_COUNTY_FIPS.get(county_code)
        if not county_fips:
            continue

        for year_label, total in info['years'].items():
            if year_label not in projected_years:
                continue
            fall_year = int(year_label.split('-')[0])
            records.append({
                'state_fips': '19',
                'state_name': 'Iowa',
                'state_abbr': 'IA',
                'county_fips': county_fips,
                'geo_name': f"{info['name']} County, IA",
                'geo_level': 'county',
                'year': fall_year,
                'projected_enrollment': round(total),
                'source': 'Iowa DOE Certified Enrollment Projections 2025-26 to 2029-30',
                'interpolated': False,
                'extrapolated': False,
            })

    counties_with_data = len(set(r['county_fips'] for r in records))
    print(f"  Iowa: {len(records)} records for {counties_with_data} counties")
    return records


def process_maryland():
    """Process Maryland county-level enrollment projections."""
    print("Processing Maryland enrollment projections...")

    xlsx_path = os.path.join(DATA_DIR, 'md_enrollment_projections.xlsx')
    if not os.path.exists(xlsx_path):
        print("  Maryland file not found, skipping")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Row 3: headers with years 2024-2034 in columns 3-13
    # Rows 4+: jurisdiction data (skip region subtotals and state total)
    records = []
    for r in range(4, 42):
        name = ws.cell(row=r, column=2).value
        if not name or not name.strip():
            continue

        name = name.strip()
        # Skip region subtotals and state total
        if name in ('MARYLAND', 'Baltimore Region', 'Washington Suburban Region',
                     'Southern Maryland', 'Western Maryland', 'Upper Eastern Shore',
                     'Upper Eastern Shore ', 'Lower Eastern Shore', 'Year'):
            continue

        county_fips = MD_COUNTY_FIPS.get(name)
        if not county_fips:
            print(f"  Warning: no FIPS for MD jurisdiction '{name}'")
            continue

        for c in range(3, 14):
            year = ws.cell(row=3, column=c).value
            val = ws.cell(row=r, column=c).value
            if year and val and isinstance(year, int):
                is_projected = year >= 2025
                records.append({
                    'state_fips': '24',
                    'state_name': 'Maryland',
                    'state_abbr': 'MD',
                    'county_fips': county_fips,
                    'geo_name': f"{name}, MD",
                    'geo_level': 'county',
                    'year': year,
                    'projected_enrollment': int(val),
                    'source': 'Maryland Dept of Planning, School Enrollment Projections 2025',
                    'interpolated': False,
                    'extrapolated': False,
                })

    counties_with_data = len(set(r['county_fips'] for r in records))
    print(f"  Maryland: {len(records)} records for {counties_with_data} counties")
    return records


def process_pennsylvania():
    """Process Pennsylvania district-level enrollment projections, aggregated to county."""
    print("Processing Pennsylvania enrollment projections...")

    xlsx_path = os.path.join(DATA_DIR, 'pa_enrollment_projections.xlsx')
    if not os.path.exists(xlsx_path):
        print("  Pennsylvania file not found, skipping")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Enrollment Projection Data']

    # Row 1: headers: Datatype, AUN, School Year, LEA Name, County, K, 001-012
    # Aggregate district projections to county level
    # county_data[county_name] = {year_label: total_enrollment}
    county_data = {}
    for r in range(2, ws.max_row + 1):
        datatype = ws.cell(row=r, column=1).value
        if datatype != 'Projection':
            continue

        county_name = ws.cell(row=r, column=5).value
        year_label = ws.cell(row=r, column=3).value  # e.g. '2025 - 2026'
        if not county_name or not year_label:
            continue

        # Sum grades K-12 (columns 6-18)
        total = 0
        for c in range(6, 19):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                total += val

        if county_name not in county_data:
            county_data[county_name] = {}
        if year_label not in county_data[county_name]:
            county_data[county_name][year_label] = 0
        county_data[county_name][year_label] += total

    # Build records
    records = []
    for county_name, years in county_data.items():
        county_fips = PA_COUNTY_FIPS.get(county_name)
        if not county_fips:
            print(f"  Warning: no FIPS for PA county '{county_name}'")
            continue

        for year_label, total in years.items():
            # Parse '2025 - 2026' -> fall year 2025
            fall_year = int(year_label.split(' - ')[0])
            records.append({
                'state_fips': '42',
                'state_name': 'Pennsylvania',
                'state_abbr': 'PA',
                'county_fips': county_fips,
                'geo_name': f"{county_name} County, PA",
                'geo_level': 'county',
                'year': fall_year,
                'projected_enrollment': round(total),
                'source': 'Pennsylvania DOE District Enrollment Projections 2024-25',
                'interpolated': False,
                'extrapolated': False,
            })

    counties_with_data = len(set(r['county_fips'] for r in records))
    print(f"  Pennsylvania: {len(records)} records for {counties_with_data} counties")
    return records


def process_virginia():
    """Process Virginia Weldon Cooper Center enrollment projections."""
    print("Processing Virginia Cooper Center enrollment projections...")
    xlsx_path = os.path.join(DATA_DIR, 'va_enrollment_projections.xlsx')
    if not os.path.exists(xlsx_path):
        print("  Skipping Virginia (file not found)")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # VA FIPS lookup for divisions (independent cities + counties)
    # We'll aggregate divisions to state level since divisions don't map cleanly to counties
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find the header row with 'Division Name'
        header_row = None
        for r in range(1, 10):
            if ws.cell(r, 1).value and 'Division' in str(ws.cell(r, 1).value):
                header_row = r
                break
        if not header_row:
            continue

        # Find the year from sheet name or column
        year_str = ws.cell(header_row, 3).value  # 'School Year' column
        if not year_str:
            # Try sheet name
            parts = sheet_name.split('-')
            if len(parts) == 2 and parts[0].strip().isdigit():
                year_str = sheet_name.strip()

        # Get the state total from the 'Virginia' row
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name and str(name).strip() == 'Virginia':
                total_col = ws.max_column  # Last column is Total
                total = ws.cell(r, total_col).value
                if total:
                    # Parse school year to calendar year
                    sy = str(ws.cell(r, 3).value or sheet_name).strip()
                    if '-' in sy:
                        cal_year = int(sy.split('-')[0])
                    elif '(' in sheet_name:
                        cal_year = int(sheet_name.split('(')[0].strip().split('-')[0])
                    else:
                        try:
                            cal_year = int(sy[:4])
                        except ValueError:
                            continue

                    is_actual = 'Actual' in sheet_name
                    records.append({
                        'state_fips': '51',
                        'state_name': 'Virginia',
                        'state_abbr': 'VA',
                        'county_fips': None,
                        'geo_name': 'Virginia',
                        'geo_level': 'state',
                        'year': cal_year,
                        'projected_enrollment': round(float(total)),
                        'source': 'Weldon Cooper Center School Enrollment Projections 2024',
                        'interpolated': False,
                        'extrapolated': False,
                    })
                break

    print(f"  Virginia: {len(records)} records (state level)")
    return records


def process_texas():
    """Process Texas TEA attendance projections."""
    print("Processing Texas TEA attendance projections...")
    xlsx_path = os.path.join(DATA_DIR, 'tx_attendance_projections.xlsx')
    if not os.path.exists(xlsx_path):
        print("  Skipping Texas (file not found)")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Aggregate district ADA to state total per year
    year_totals = {}
    for r in range(2, ws.max_row + 1):
        yr_str = ws.cell(r, 1).value
        ada = ws.cell(r, 4).value  # ADA column
        if yr_str and ada:
            yr = int(yr_str)
            year_totals[yr] = year_totals.get(yr, 0) + float(ada)

    records = []
    for yr in sorted(year_totals.keys()):
        records.append({
            'state_fips': '48',
            'state_name': 'Texas',
            'state_abbr': 'TX',
            'county_fips': None,
            'geo_name': 'Texas',
            'geo_level': 'state',
            'year': yr,
            'projected_enrollment': round(year_totals[yr]),
            'source': 'Texas Education Agency Attendance Projections FY2026-2027',
            'interpolated': False,
            'extrapolated': False,
        })

    print(f"  Texas: {len(records)} records (state level, {sorted(year_totals.keys())})")
    return records


def process_colorado_population():
    """Process Colorado DOLA county population by age -> school-age proxy."""
    print("Processing Colorado DOLA school-age population projections...")
    import csv
    csv_path = os.path.join(DATA_DIR, 'co_county_population_by_age.csv')
    if not os.path.exists(csv_path):
        print("  Skipping Colorado (file not found)")
        return []

    # Read CSV, skip vintage header line
    with open(csv_path, 'r') as f:
        lines = f.readlines()

    import io
    reader = csv.DictReader(io.StringIO(''.join(lines[1:])))

    # Each row is one county/year/age combination with totalpopulation
    # Sum ages 5-17 by year across all counties
    year_totals = {}
    for row in reader:
        yr = int(row.get('year', 0))
        if yr < 2022 or yr > 2050:
            continue
        age_val = int(row.get('age', -1))
        if age_val < 5 or age_val > 17:
            continue
        pop = int(row.get('totalpopulation', 0))
        year_totals[yr] = year_totals.get(yr, 0) + pop

    # Convert population to enrollment using 2022 ratio
    # CO 2022 NCES enrollment: ~870,900; 2022 school-age pop from our data
    pop_2022 = year_totals.get(2022, 1)
    nces_2022 = 870900  # from NCES
    enroll_ratio = nces_2022 / pop_2022 if pop_2022 > 0 else 0.9

    records = []
    for yr in sorted(year_totals.keys()):
        records.append({
            'state_fips': '08',
            'state_name': 'Colorado',
            'state_abbr': 'CO',
            'county_fips': None,
            'geo_name': 'Colorado',
            'geo_level': 'state',
            'year': yr,
            'projected_enrollment': round(year_totals[yr] * enroll_ratio),
            'source': 'Colorado DOLA School-Age Population Projections (ages 5-17)',
            'interpolated': False,
            'extrapolated': False,
        })

    print(f"  Colorado: {len(records)} records (state level, pop->enrollment ratio={enroll_ratio:.3f})")
    return records


def process_nc_population():
    """Process North Carolina OSBM county population by age -> school-age proxy."""
    print("Processing North Carolina OSBM school-age population projections...")
    import csv
    csv_path = os.path.join(DATA_DIR, 'nc_county_population_by_age.csv')
    if not os.path.exists(csv_path):
        print("  Skipping North Carolina (file not found)")
        return []

    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        # Sum ages 5-17 by year (state total), using only "Total" sex rows
        year_totals = {}
        for row in reader:
            yr = int(row.get('year', 0))
            if yr < 2022 or yr > 2050:
                continue
            if row.get('sex', '').strip() != 'Total':
                continue
            school_age = sum(int(row.get(f'age{a}', 0)) for a in range(5, 18))
            year_totals[yr] = year_totals.get(yr, 0) + school_age

    # Convert to enrollment using 2022 ratio
    pop_2022 = year_totals.get(2022, 1)
    nces_2022 = 1531800  # from NCES
    enroll_ratio = nces_2022 / pop_2022 if pop_2022 > 0 else 0.9

    records = []
    for yr in sorted(year_totals.keys()):
        records.append({
            'state_fips': '37',
            'state_name': 'North Carolina',
            'state_abbr': 'NC',
            'county_fips': None,
            'geo_name': 'North Carolina',
            'geo_level': 'state',
            'year': yr,
            'projected_enrollment': round(year_totals[yr] * enroll_ratio),
            'source': 'NC OSBM School-Age Population Projections (ages 5-17)',
            'interpolated': False,
            'extrapolated': False,
        })

    print(f"  North Carolina: {len(records)} records (state level, pop->enrollment ratio={enroll_ratio:.3f})")
    return records


def load_acs_school_age():
    """Load ACS county school-age population shares for distributing state projections."""
    acs_path = os.path.join(DATA_DIR, 'acs_county_school_age.json')
    if not os.path.exists(acs_path):
        print("  ACS school-age data not found")
        return {}

    with open(acs_path, 'r') as f:
        acs = json.load(f)

    # Build state -> {county_fips: share} mapping
    # ACS data is dict keyed by county FIPS with {name, state_fips, pop_total, school_age}
    state_counties = {}
    if isinstance(acs, dict):
        for county_fips, info in acs.items():
            state_fips = info.get('state_fips', county_fips[:2])
            school_age = info.get('school_age', 0)
            if state_fips not in state_counties:
                state_counties[state_fips] = {}
            state_counties[state_fips][county_fips] = school_age
    else:
        for county in acs:
            state_fips = county.get('state_fips', county.get('state'))
            county_fips = county.get('county_fips', county.get('fips'))
            school_age = county.get('school_age', county.get('pop_5_17', 0))
            if not state_fips or not county_fips:
                continue
            if state_fips not in state_counties:
                state_counties[state_fips] = {}
            state_counties[state_fips][county_fips] = school_age

    # Convert to shares
    state_shares = {}
    for state_fips, counties in state_counties.items():
        total = sum(counties.values())
        if total > 0:
            state_shares[state_fips] = {
                cfips: pop / total for cfips, pop in counties.items()
            }

    # Build county names lookup
    county_names = {}
    if isinstance(acs, dict):
        for county_fips, info in acs.items():
            county_names[county_fips] = info.get('name', county_fips)

    return state_shares, county_names


def distribute_to_counties(nces_records, state_shares, county_names=None):
    """For states without county-level projections, distribute state totals to counties."""
    print("Distributing state projections to counties using ACS shares...")

    # Skip states with their own county-level data (CA, IA, MD, PA)
    skip_states = {'06', '19', '24', '42'}

    county_records = []
    for rec in nces_records:
        sfips = rec['state_fips']
        if sfips in skip_states:
            continue
        if sfips not in state_shares:
            continue

        for county_fips, share in state_shares[sfips].items():
            geo_name = county_names.get(county_fips, county_fips) if county_names else county_fips
            county_records.append({
                'state_fips': sfips,
                'state_name': rec['state_name'],
                'state_abbr': rec['state_abbr'],
                'county_fips': county_fips,
                'geo_name': geo_name,
                'geo_level': 'county_distributed',
                'year': rec['year'],
                'projected_enrollment': round(rec['projected_enrollment'] * share),
                'source': f"NCES state projection distributed to county via ACS school-age shares",
                'interpolated': rec.get('interpolated', False),
                'extrapolated': rec.get('extrapolated', False),
            })

    print(f"  Distributed: {len(county_records)} county-level records")
    return county_records


def build_coverage_report(state_records, county_records):
    """Build data_coverage_report.md."""
    print("Building coverage report...")

    # Load sources inventory
    with open(os.path.join(OUT_DIR, 'state_sources.json'), 'r') as f:
        sources = json.load(f)

    lines = [
        "# State Enrollment Projections — Data Coverage Report",
        "",
        f"Generated: 2026-02-13",
        "",
        "## Summary",
        "",
    ]

    # Categorize states
    downloaded = []
    available = []
    pdf_only = []
    dashboard_only = []
    contract_only = []
    no_proj = []
    historical_only = []

    for s in sources['states']:
        status = s.get('status', '')
        if status == 'downloaded':
            downloaded.append(s)
        elif status == 'available':
            available.append(s)
        elif 'PDF' in status:
            pdf_only.append(s)
        elif 'dashboard' in status:
            dashboard_only.append(s)
        elif 'contract' in status:
            contract_only.append(s)
        elif 'historical' in status:
            historical_only.append(s)
        elif 'population' in status:
            no_proj.append(s)
        else:
            no_proj.append(s)

    lines.append(f"| Category | Count | States |")
    lines.append(f"|----------|-------|--------|")
    lines.append(f"| Downloaded & processed | {len(downloaded)} | {', '.join(s['abbr'] for s in downloaded)} |")
    lines.append(f"| Available (downloadable Excel) | {len(available)} | {', '.join(s['abbr'] for s in available)} |")
    lines.append(f"| PDF — manual extraction needed | {len(pdf_only)} | {', '.join(s['abbr'] for s in pdf_only)} |")
    lines.append(f"| Dashboard — no bulk download | {len(dashboard_only)} | {', '.join(s['abbr'] for s in dashboard_only)} |")
    lines.append(f"| Contract/paid basis only | {len(contract_only)} | {', '.join(s['abbr'] for s in contract_only)} |")
    lines.append(f"| Historical enrollment only | {len(historical_only)} | {', '.join(s['abbr'] for s in historical_only)} |")
    lines.append(f"| No projections found | {len(no_proj)} | {', '.join(s['abbr'] for s in no_proj)} |")
    lines.append("")

    lines.append("## NCES Baseline Coverage")
    lines.append("")
    lines.append("All 50 states + DC are covered at the **state level** by NCES Projections of Education Statistics")
    lines.append("to 2031 (Digest Table 203.20). Years 2026-2030 are linearly interpolated; 2032-2050 are")
    lines.append("linearly extrapolated from the 2025-2031 trend.")
    lines.append("")
    lines.append(f"- **State-level records**: {len(state_records)}")
    lines.append(f"- **County-level records** (CA, IA, MD, PA direct + ACS-distributed): {len(county_records)}")
    lines.append("")

    lines.append("## State-by-State Detail")
    lines.append("")
    lines.append("| State | FIPS | Source | Geo Level | Horizon | Format | Status |")
    lines.append("|-------|------|--------|-----------|---------|--------|--------|")
    for s in sources['states']:
        lines.append(f"| {s['abbr']} | {s['fips']} | {s['source'][:40]} | {s['geo_level']} | {s.get('time_horizon', 'N/A')[:25]} | {s.get('format', 'N/A')} | {s['status']} |")
    lines.append("")

    lines.append("## Downloaded & Processed Sources")
    lines.append("")
    lines.append("### 1. NCES Projections of Education Statistics to 2031")
    lines.append("- **Table**: Digest Table 203.20")
    lines.append("- **Coverage**: All 50 states + DC, state level")
    lines.append("- **Years**: Fall 2022, 2023, 2024, 2025, 2031 (actual); 2026-2030 interpolated; 2032-2050 extrapolated")
    lines.append("- **Method**: NCES National Elementary and Secondary Enrollment Projection Model")
    lines.append("")
    lines.append("### 2. California Department of Finance, 2025 Series")
    lines.append("- **Coverage**: 58 California counties")
    lines.append("- **Years**: 2024-25 through 2046-47 (22 projection years)")
    lines.append("- **Method**: Grade progression ratio (cohort-survival) model")
    lines.append("- **File**: `data/ca_k12_enrollment_2025.xlsx`")
    lines.append("")
    lines.append("### 3. Iowa Department of Education, Certified Enrollment Projections")
    lines.append("- **Coverage**: 99 Iowa counties (aggregated from ~330 districts)")
    lines.append("- **Years**: 2025-26 through 2029-30 (5 projection years)")
    lines.append("- **Method**: Grade progression rate + birth data")
    lines.append("- **File**: `data/ia_enrollment_projections.xlsx`")
    lines.append("")
    lines.append("### 4. Maryland Department of Planning, School Enrollment Projections 2025")
    lines.append("- **Coverage**: 24 Maryland jurisdictions (23 counties + Baltimore City)")
    lines.append("- **Years**: 2024 actual + 2025-2034 projected (10 projection years)")
    lines.append("- **Method**: Population projections + grade succession methodology")
    lines.append("- **File**: `data/md_enrollment_projections.xlsx`")
    lines.append("")
    lines.append("### 5. Pennsylvania Department of Education, District Enrollment Projections")
    lines.append("- **Coverage**: 67 Pennsylvania counties (aggregated from 499 districts)")
    lines.append("- **Years**: 2025-26 through 2034-35 (10 projection years)")
    lines.append("- **Method**: Grade progression + modified enrollment rate model")
    lines.append("- **File**: `data/pa_enrollment_projections.xlsx`")
    lines.append("")

    lines.append("## Gaps and Next Steps")
    lines.append("")
    lines.append("### States with Downloadable Projections Not Yet Processed")
    for s in available:
        lines.append(f"- **{s['state']}** ({s['abbr']}): {s['source']} — {s.get('format', '?')} at {s['geo_level']} level")
    lines.append("")

    lines.append("### States with PDF-Only Projections")
    for s in pdf_only:
        lines.append(f"- **{s['state']}** ({s['abbr']}): {s['source']} — {s.get('notes', '')[:80]}")
    lines.append("")

    lines.append("### States with No Published Projections")
    lines.append("These states rely on NCES state-level projections distributed to counties via ACS school-age population shares:")
    for s in no_proj + historical_only:
        lines.append(f"- {s['state']} ({s['abbr']})")
    lines.append("")

    report = '\n'.join(lines)
    report_path = os.path.join(OUT_DIR, 'data_coverage_report.md')
    with open(report_path, 'w') as f:
        f.write(report)
    print(f"  Coverage report written to {report_path}")


def main():
    print("=" * 60)
    print("State Enrollment Projections Processing")
    print("=" * 60)

    # Step 1: Process NCES state-level data
    nces_records = process_nces()

    # Step 2: Process state-specific data
    ca_records = process_california()
    ia_records = process_iowa()
    md_records = process_maryland()
    pa_records = process_pennsylvania()
    va_records = process_virginia()
    tx_records = process_texas()
    co_records = process_colorado_population()
    nc_records = process_nc_population()
    state_specific = ca_records + ia_records + md_records + pa_records + va_records + tx_records + co_records + nc_records

    # Step 3: Load ACS shares and distribute state projections to counties
    result = load_acs_school_age()
    if result:
        state_shares, county_names = result
        county_distributed = distribute_to_counties(nces_records, state_shares, county_names)
    else:
        county_distributed = []
        print("  Skipping county distribution (no ACS data)")

    # Combine all records
    all_records = nces_records + state_specific + county_distributed

    # Build output JSON
    output = {
        'metadata': {
            'description': 'Combined state and county enrollment projections from multiple sources',
            'sources': [
                'NCES Projections of Education Statistics to 2031 (all states, state level)',
                'California DOF K-12 Enrollment Projections 2025 Series (CA counties)',
                'Iowa DOE Certified Enrollment Projections 2025-26 to 2029-30 (IA counties)',
                'Maryland Dept of Planning School Enrollment Projections 2025 (MD counties)',
                'Pennsylvania DOE District Enrollment Projections 2024-25 (PA counties)',
                'Weldon Cooper Center School Enrollment Projections 2024 (VA state level)',
                'Texas Education Agency Attendance Projections FY2026-2027 (TX state level)',
                'Colorado DOLA School-Age Population Projections (CO state level, ages 5-17)',
                'NC OSBM School-Age Population Projections (NC state level, ages 5-17)',
            ],
            'total_records': len(all_records),
            'state_level_records': len(nces_records),
            'county_level_records': len(state_specific) + len(county_distributed),
            'years_covered': '2022-2050',
            'notes': 'NCES years 2026-2030 interpolated, 2032-2050 extrapolated. Non-CA counties distributed from state totals using ACS school-age population shares.',
        },
        'projections': all_records,
    }

    out_path = os.path.join(OUT_DIR, 'state_enrollment_projections.json')
    with open(out_path, 'w') as f:
        json.dump(output, f)
    print(f"\nOutput: {out_path} ({os.path.getsize(out_path) / 1024 / 1024:.1f} MB)")
    print(f"Total records: {len(all_records)}")

    # Build coverage report
    build_coverage_report(nces_records, state_specific + county_distributed)

    print("\nDone!")


if __name__ == '__main__':
    main()
