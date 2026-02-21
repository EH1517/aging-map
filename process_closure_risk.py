"""
State-Level School Closure Risk Map — Data Processing
=====================================================
Downloads and processes source data into JSON files for the interactive map.

Sources:
- CDC NCHS state fertility rates (2022) via USAFacts (hardcoded)
- Census Bureau 2023 National Population Projections (np2023-t2.xlsx)
- NCES Common Core of Data 2022-23 school directory/membership
- Census TIGER/Line state boundaries

Outputs:
- state_projections.json — projections by state, scenario, year
- states.json — state boundary GeoJSON
"""

import json
import os
import sys
import math
import zipfile
import shutil

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# A. Hardcoded state fertility data (CDC NCHS via USAFacts)
# ============================================================
# rate_2022: births per 1,000 women ages 15-44
# decline_2005_2022: proportional decline (negative = decline)

STATE_FERTILITY = {
    "South Dakota":    {"rate_2022": 66.5, "decline_2005_2022": -0.110},
    "Alaska":          {"rate_2022": 64.9, "decline_2005_2022": -0.122},
    "Nebraska":        {"rate_2022": 63.6, "decline_2005_2022": -0.130},
    "North Dakota":    {"rate_2022": 62.0, "decline_2005_2022": -0.043},
    "Texas":           {"rate_2022": 61.9, "decline_2005_2022": -0.200},
    "Louisiana":       {"rate_2022": 61.8, "decline_2005_2022": -0.003},
    "Utah":            {"rate_2022": 61.3, "decline_2005_2022": -0.339},
    "Kentucky":        {"rate_2022": 61.1, "decline_2005_2022": -0.060},
    "Oklahoma":        {"rate_2022": 60.4, "decline_2005_2022": -0.158},
    "Kansas":          {"rate_2022": 60.3, "decline_2005_2022": -0.154},
    "Arkansas":        {"rate_2022": 60.2, "decline_2005_2022": -0.134},
    "Iowa":            {"rate_2022": 59.9, "decline_2005_2022": -0.105},
    "Indiana":         {"rate_2022": 59.7, "decline_2005_2022": -0.109},
    "Mississippi":     {"rate_2022": 59.7, "decline_2005_2022": -0.131},
    "Hawaii":          {"rate_2022": 59.3, "decline_2005_2022": -0.143},
    "Tennessee":       {"rate_2022": 59.3, "decline_2005_2022": -0.091},
    "Alabama":         {"rate_2022": 58.7, "decline_2005_2022": -0.077},
    "New Jersey":      {"rate_2022": 58.7, "decline_2005_2022": -0.079},
    "Idaho":           {"rate_2022": 58.4, "decline_2005_2022": -0.260},
    "Minnesota":       {"rate_2022": 58.2, "decline_2005_2022": -0.122},
    "Missouri":        {"rate_2022": 57.7, "decline_2005_2022": -0.124},
    "North Carolina":  {"rate_2022": 57.6, "decline_2005_2022": -0.135},
    "District of Columbia": {"rate_2022": 57.3, "decline_2005_2022": -0.145},
    "Ohio":            {"rate_2022": 57.3, "decline_2005_2022": -0.098},
    "South Carolina":  {"rate_2022": 57.0, "decline_2005_2022": -0.114},
    "Maryland":        {"rate_2022": 56.9, "decline_2005_2022": -0.082},
    "Georgia":         {"rate_2022": 56.0, "decline_2005_2022": -0.216},
    "Florida":         {"rate_2022": 55.6, "decline_2005_2022": -0.135},
    "Virginia":        {"rate_2022": 55.6, "decline_2005_2022": -0.131},
    "Wyoming":         {"rate_2022": 55.4, "decline_2005_2022": -0.224},
    "Arizona":         {"rate_2022": 54.9, "decline_2005_2022": -0.319},
    "Wisconsin":       {"rate_2022": 54.2, "decline_2005_2022": -0.133},
    "Michigan":        {"rate_2022": 54.0, "decline_2005_2022": -0.128},
    "West Virginia":   {"rate_2022": 54.0, "decline_2005_2022": -0.088},
    "New York":        {"rate_2022": 53.6, "decline_2005_2022": -0.105},
    "Pennsylvania":    {"rate_2022": 53.3, "decline_2005_2022": -0.089},
    "Washington":      {"rate_2022": 53.3, "decline_2005_2022": -0.154},
    "Montana":         {"rate_2022": 53.2, "decline_2005_2022": -0.169},
    "Nevada":          {"rate_2022": 53.2, "decline_2005_2022": -0.280},
    "New Mexico":      {"rate_2022": 53.1, "decline_2005_2022": -0.273},
    "California":      {"rate_2022": 52.8, "decline_2005_2022": -0.251},
    "Illinois":        {"rate_2022": 51.8, "decline_2005_2022": -0.227},
    "Colorado":        {"rate_2022": 51.5, "decline_2005_2022": -0.260},
    "Connecticut":     {"rate_2022": 50.7, "decline_2005_2022": -0.138},
    "Maine":           {"rate_2022": 49.7, "decline_2005_2022": -0.085},
    "Massachusetts":   {"rate_2022": 48.7, "decline_2005_2022": -0.129},
    "New Hampshire":   {"rate_2022": 47.9, "decline_2005_2022": -0.115},
    "Rhode Island":    {"rate_2022": 47.5, "decline_2005_2022": -0.146},
    "Oregon":          {"rate_2022": 47.3, "decline_2005_2022": -0.248},
    "Delaware":        {"rate_2022": 44.9, "decline_2005_2022": -0.312},
    "Vermont":         {"rate_2022": 44.3, "decline_2005_2022": -0.118},
}

STATE_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}

STATE_FIPS = {
    "Alabama": "01", "Alaska": "02", "Arizona": "04", "Arkansas": "05",
    "California": "06", "Colorado": "08", "Connecticut": "09", "Delaware": "10",
    "District of Columbia": "11", "Florida": "12", "Georgia": "13", "Hawaii": "15",
    "Idaho": "16", "Illinois": "17", "Indiana": "18", "Iowa": "19",
    "Kansas": "20", "Kentucky": "21", "Louisiana": "22", "Maine": "23",
    "Maryland": "24", "Massachusetts": "25", "Michigan": "26", "Minnesota": "27",
    "Mississippi": "28", "Missouri": "29", "Montana": "30", "Nebraska": "31",
    "Nevada": "32", "New Hampshire": "33", "New Jersey": "34", "New Mexico": "35",
    "New York": "36", "North Carolina": "37", "North Dakota": "38", "Ohio": "39",
    "Oklahoma": "40", "Oregon": "41", "Pennsylvania": "42", "Rhode Island": "44",
    "South Carolina": "45", "South Dakota": "46", "Tennessee": "47", "Texas": "48",
    "Utah": "49", "Vermont": "50", "Virginia": "51", "Washington": "53",
    "West Virginia": "54", "Wisconsin": "55", "Wyoming": "56",
}

PROJECTION_YEARS = [2025, 2027, 2030, 2033, 2035, 2037, 2040]

FERTILITY_FLOOR = 35.0  # births per 1,000 women 15-44

# Annual closure probabilities by condition.
# Recalibrated against Brookings (2024) historical closure rate data:
# most states <0.5%/yr historically; highest observed = DC at 2.06%/yr.
# Forward-looking rates are set above historical averages to reflect
# worsening enrollment conditions, but scaled to remain defensible.
CLOSURE_PROBS = {
    'large_stable':   0.0075,  # background rate: <=3% decline
    'mild_decline':   0.010,   # 3-10% decline
    'medium_decline': 0.015,   # >10% decline
    'small_steep':    0.03,    # 100-199 projected enr with >25% decline
    'tiny':           0.04,    # <100 projected enrollment
    'default':        0.0075,  # fallback
}


def download_file(url, dest, desc="file"):
    """Download a file if it doesn't exist."""
    if os.path.exists(dest):
        print(f"  {desc}: cached at {os.path.basename(dest)}")
        return True

    print(f"  Downloading {desc}...")
    import urllib.request
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        resp = urllib.request.urlopen(req, timeout=60)
        with open(dest, 'wb') as f:
            f.write(resp.read())
        print(f"  Downloaded {os.path.basename(dest)} ({os.path.getsize(dest) / 1024:.0f} KB)")
        return True
    except Exception as e:
        print(f"  ERROR downloading {desc}: {e}")
        return False


# ============================================================
# Step 1: Parse CCD school data -> state school counts
# ============================================================

def fetch_ccd_data():
    """Fetch CCD school data from Urban Institute API or use cached."""
    cache_path = os.path.join(DATA_DIR, 'ccd_schools_2022.json')

    # Try copying from aging-map project first
    aging_map_path = os.path.join(os.path.dirname(OUT_DIR), 'aging-map', 'data', 'ccd_schools_2022.json')
    if not os.path.exists(cache_path) and os.path.exists(aging_map_path):
        print("  Copying CCD data from aging-map project...")
        shutil.copy2(aging_map_path, cache_path)

    if os.path.exists(cache_path):
        print(f"  CCD data: cached ({os.path.getsize(cache_path) / 1024 / 1024:.1f} MB)")
        with open(cache_path) as f:
            return json.load(f)

    # Fetch from Urban Institute Education Data Portal API
    print("  Fetching CCD school data from Urban Institute API...")
    import urllib.request
    import time

    all_schools = []
    page = 0
    per_page = 10000
    while True:
        url = (f"https://educationdata.urban.org/api/v1/schools/ccd/directory/2022/"
               f"?school_type=1&limit={per_page}&offset={page * per_page}")
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            resp = urllib.request.urlopen(req, timeout=60)
            data = json.loads(resp.read())
            results = data.get('results', [])
            if not results:
                break
            all_schools.extend(results)
            print(f"    Page {page}: {len(results)} schools (total: {len(all_schools)})")
            page += 1
            time.sleep(1)
        except Exception as e:
            print(f"    Error on page {page}: {e}")
            if page > 0:
                time.sleep(3)
                continue
            break

    if all_schools:
        with open(cache_path, 'w') as f:
            json.dump(all_schools, f)
        print(f"  Cached {len(all_schools)} schools")

    return all_schools


def process_ccd_schools():
    """Process CCD data into state-level school counts by size bucket."""
    print("\n[Step 1] Processing CCD school data...")

    schools = fetch_ccd_data()
    if not schools:
        print("  ERROR: No CCD data available")
        sys.exit(1)

    # Filter to regular, operational, non-charter, non-virtual schools with enrollment
    filtered = []
    for s in schools:
        school_type = s.get('school_type')
        charter = s.get('charter')
        virtual = s.get('virtual')
        enrollment = s.get('enrollment')
        school_status = s.get('school_status')

        # Type 1 = regular school
        if school_type != 1:
            continue
        # Exclude charters
        if charter == 1:
            continue
        # Exclude virtual
        if virtual and virtual in (1, 'Yes'):
            continue
        # Must have enrollment > 0
        if not enrollment or enrollment <= 0:
            continue
        # Must be operational (status 1 or 3 = open)
        if school_status and school_status not in (1, 3):
            continue

        filtered.append(s)

    print(f"  Filtered to {len(filtered)} regular, operational TPS schools")

    # Aggregate to state level
    # state_data[state_abbr] = {total_schools, total_enrollment, buckets: {<100, 100-199, 200-299, 300-499, 500+}}
    state_data = {}
    for s in filtered:
        st = s.get('state_location') or s.get('state_mailing')
        if not st or len(st) != 2:
            continue

        enrollment = s['enrollment']

        if st not in state_data:
            state_data[st] = {
                'total_schools': 0,
                'total_enrollment': 0,
                'bucket_under100': 0,
                'bucket_100_199': 0,
                'bucket_200_299': 0,
                'bucket_300_499': 0,
                'bucket_500plus': 0,
            }

        sd = state_data[st]
        sd['total_schools'] += 1
        sd['total_enrollment'] += enrollment

        if enrollment < 100:
            sd['bucket_under100'] += 1
        elif enrollment < 200:
            sd['bucket_100_199'] += 1
        elif enrollment < 300:
            sd['bucket_200_299'] += 1
        elif enrollment < 500:
            sd['bucket_300_499'] += 1
        else:
            sd['bucket_500plus'] += 1

    print(f"  States with data: {len(state_data)}")
    total_schools = sum(sd['total_schools'] for sd in state_data.values())
    total_enrollment = sum(sd['total_enrollment'] for sd in state_data.values())
    print(f"  Total schools: {total_schools:,}")
    print(f"  Total enrollment: {total_enrollment:,}")

    return state_data


# ============================================================
# Step 2: Extract Census national birth trajectory
# ============================================================

def extract_census_birth_trajectory():
    """Extract Under-5 population trajectory from Census np2023-t2.xlsx as birth proxy."""
    print("\n[Step 2] Extracting Census Bureau birth trajectory...")

    xlsx_path = os.path.join(DATA_DIR, 'np2023-t2.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"  ERROR: {xlsx_path} not found")
        print("  Download from: https://www.census.gov/data/tables/2023/demo/popproj/2023-summary-tables.html")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Main series (thousands)']

    # Row 6: year headers (2022, 2025, 2030, 2035, 2040, ...)
    # Row 9: Under 5 years (proxy for births ~5 years earlier)
    years = []
    under5 = []
    for c in range(2, 19):
        year = ws.cell(row=6, column=c).value
        val = ws.cell(row=9, column=c).value
        if year and val:
            years.append(int(year))
            under5.append(float(val))

    # Compute decline multipliers relative to 2022 baseline
    baseline = under5[0]  # 2022 value
    multipliers = {}
    for i, year in enumerate(years):
        multipliers[year] = under5[i] / baseline

    # Interpolate to get annual values
    annual_multipliers = {}
    for y in range(2022, 2051):
        if y in multipliers:
            annual_multipliers[y] = multipliers[y]
        else:
            # Linear interpolation between surrounding known years
            lower = max(yr for yr in years if yr <= y)
            upper = min(yr for yr in years if yr >= y)
            if lower == upper:
                annual_multipliers[y] = multipliers[lower]
            else:
                frac = (y - lower) / (upper - lower)
                annual_multipliers[y] = multipliers[lower] + frac * (multipliers[upper] - multipliers[lower])

    print(f"  Census Under-5 trajectory: {years[0]}={under5[0]:.0f}K -> {years[-1]}={under5[-1]:.0f}K")
    print(f"  2022->2030 multiplier: {annual_multipliers.get(2030, 0):.4f}")
    print(f"  2022->2040 multiplier: {annual_multipliers.get(2040, 0):.4f}")

    return annual_multipliers


# ============================================================
# Step 3: Build fertility projections (both scenarios)
# ============================================================

def build_fertility_projections(census_multipliers):
    """Build two fertility scenarios for each state."""
    print("\n[Step 3] Building fertility projections...")

    projections = {}

    for state_name, fdata in STATE_FERTILITY.items():
        abbr = STATE_ABBR[state_name]
        rate_2022 = fdata['rate_2022']
        decline_17yr = abs(fdata['decline_2005_2022'])

        # Annualized compound decline rate for state trend scenario
        # decline over 17 years: rate_2005 * (1 - annual_rate)^17 = rate_2022
        # (1 - annual_rate)^17 = 1 - decline_17yr
        # annual_rate = 1 - (1 - decline_17yr)^(1/17)
        if decline_17yr > 0:
            annual_decline = 1.0 - (1.0 - decline_17yr) ** (1.0 / 17.0)
        else:
            annual_decline = 0.0

        state_proj = {
            'state': state_name,
            'abbr': abbr,
            'fips': STATE_FIPS[state_name],
            'rate_2022': rate_2022,
            'decline_2005_2022': fdata['decline_2005_2022'],
            'annual_decline_rate': round(annual_decline, 6),
            'census': {},  # scenario 1
            'state_trend': {},  # scenario 2
        }

        for year in range(2022, 2051):
            # Scenario 1: Census Bureau uniform national curve
            mult = census_multipliers.get(year, 1.0)
            census_rate = rate_2022 * mult
            state_proj['census'][year] = {
                'rate': round(census_rate, 2),
                'birth_decline_pct': round((1.0 - mult) * 100, 2),
            }

            # Scenario 2: State trend extrapolation
            years_from_2022 = year - 2022
            trend_rate = rate_2022 * ((1.0 - annual_decline) ** years_from_2022)
            trend_rate = max(trend_rate, FERTILITY_FLOOR)
            trend_mult = trend_rate / rate_2022
            state_proj['state_trend'][year] = {
                'rate': round(trend_rate, 2),
                'birth_decline_pct': round((1.0 - trend_mult) * 100, 2),
            }

        projections[abbr] = state_proj

    # Print some examples
    for abbr in ['UT', 'LA', 'CA', 'OH']:
        p = projections.get(abbr, {})
        if p:
            c2040 = p['census'].get(2040, {})
            s2040 = p['state_trend'].get(2040, {})
            print(f"  {abbr}: rate_2022={p['rate_2022']}, "
                  f"census_2040={c2040.get('rate', '?')}, "
                  f"state_trend_2040={s2040.get('rate', '?')}")

    return projections


# ============================================================
# Step 4: Compute lagged enrollment decline
# ============================================================

def compute_enrollment_decline(fertility_projections):
    """Compute K-12 enrollment decline from birth decline with pipeline lag."""
    print("\n[Step 4] Computing lagged enrollment decline...")

    # Children born in year X enter K in X+5, exit 12th grade in X+17
    # K-12 population in year Y consists of birth cohorts from Y-17 to Y-5
    # Post-2022 cohorts: those born 2023 and later

    for abbr, proj in fertility_projections.items():
        for scenario in ['census', 'state_trend']:
            scenario_data = proj[scenario]

            for target_year in range(2025, 2051):
                # Birth cohorts currently in K-12 in target_year:
                # Born from (target_year - 17) to (target_year - 5)
                birth_year_start = target_year - 17  # oldest K-12 students (12th graders)
                birth_year_end = target_year - 5      # youngest (kindergartners)

                total_grades = 13  # K through 12
                post_2022_grades = 0
                total_birth_decline = 0.0

                for birth_year in range(birth_year_start, birth_year_end + 1):
                    if birth_year > 2022:
                        # This cohort is post-2022; apply birth decline
                        post_2022_grades += 1
                        bd = scenario_data.get(birth_year, {}).get('birth_decline_pct', 0)
                        total_birth_decline += bd

                # Average birth decline across post-2022 cohorts in pipeline
                if post_2022_grades > 0:
                    avg_birth_decline = total_birth_decline / post_2022_grades
                else:
                    avg_birth_decline = 0.0

                # Fraction of K-12 pipeline that is post-2022
                post_2022_fraction = post_2022_grades / total_grades

                # Enrollment decline = fraction of pipeline affected × average birth decline
                enrollment_decline_pct = post_2022_fraction * avg_birth_decline

                scenario_data[target_year]['post_2022_fraction'] = round(post_2022_fraction, 3)
                scenario_data[target_year]['enrollment_decline_pct'] = round(enrollment_decline_pct, 2)

    # Print examples
    for abbr in ['UT', 'LA', 'CA']:
        p = fertility_projections[abbr]
        for yr in [2030, 2035, 2040]:
            c = p['census'].get(yr, {})
            s = p['state_trend'].get(yr, {})
            print(f"  {abbr} {yr}: census enroll_decline={c.get('enrollment_decline_pct', '?')}%, "
                  f"state_trend={s.get('enrollment_decline_pct', '?')}%, "
                  f"pipeline_fraction={c.get('post_2022_fraction', '?')}")

    return fertility_projections


# ============================================================
# Step 5: Estimate school closures
# ============================================================

def estimate_closures(fertility_projections, state_schools):
    """Estimate expected school closures using evidence-based probabilities."""
    print("\n[Step 5] Estimating school closures...")

    abbr_to_name = {v: k for k, v in STATE_ABBR.items()}

    for abbr, proj in fertility_projections.items():
        schools = state_schools.get(abbr)
        if not schools:
            continue

        total = schools['total_schools']
        enrollment = schools['total_enrollment']

        # Current school size distribution
        buckets = {
            'under100': schools['bucket_under100'],
            '100_199': schools['bucket_100_199'],
            '200_299': schools['bucket_200_299'],
            '300_499': schools['bucket_300_499'],
            '500plus': schools['bucket_500plus'],
        }

        # Average enrollment per school in each bucket (approximate midpoints)
        avg_enrollment = {
            'under100': 55,
            '100_199': 150,
            '200_299': 250,
            '300_499': 400,
            '500plus': max(750, enrollment / max(buckets['500plus'], 1)) if buckets['500plus'] > 0 else 750,
        }

        proj['current_schools'] = total
        proj['current_enrollment'] = enrollment
        proj['buckets'] = buckets

        for scenario in ['census', 'state_trend']:
            scenario_data = proj[scenario]

            for target_year in PROJECTION_YEARS:
                yr_data = scenario_data.get(target_year, {})
                enroll_decline_pct = yr_data.get('enrollment_decline_pct', 0)
                decline_ratio = enroll_decline_pct / 100.0  # as fraction

                # Apply decline to each bucket's average enrollment
                # Then reclassify and assign closure probabilities
                years_from_2025 = max(target_year - 2025, 1)

                expected_closures = 0.0
                risk_elevated = 0  # schools in yellow/orange/red
                risk_severe = 0    # schools in orange/red

                for bucket_name, count in buckets.items():
                    if count == 0:
                        continue

                    avg_enr = avg_enrollment[bucket_name]
                    projected_enr = avg_enr * (1.0 - decline_ratio)

                    # Determine annual closure probability based on post-decline size and decline magnitude
                    if projected_enr < 100:
                        annual_prob = CLOSURE_PROBS['tiny']
                        risk_severe += count
                        risk_elevated += count
                    elif projected_enr < 200 and decline_ratio > 0.25:
                        annual_prob = CLOSURE_PROBS['small_steep']
                        risk_severe += count
                        risk_elevated += count
                    elif decline_ratio > 0.10:
                        annual_prob = CLOSURE_PROBS['medium_decline']
                        risk_elevated += count
                    elif decline_ratio > 0.03:
                        annual_prob = CLOSURE_PROBS['mild_decline']
                        risk_elevated += count
                    else:
                        annual_prob = CLOSURE_PROBS['large_stable']

                    # Cumulative probability over projection period
                    cum_prob = 1.0 - (1.0 - annual_prob) ** years_from_2025
                    expected_closures += count * cum_prob

                projected_enrollment = round(enrollment * (1.0 - decline_ratio))

                yr_data['projected_enrollment'] = projected_enrollment
                yr_data['expected_closures'] = round(expected_closures, 1)
                yr_data['expected_closures_pct'] = round(expected_closures / total * 100, 1) if total > 0 else 0
                yr_data['schools_elevated_risk'] = risk_elevated
                yr_data['schools_elevated_risk_pct'] = round(risk_elevated / total * 100, 1) if total > 0 else 0
                yr_data['schools_severe_risk'] = risk_severe
                yr_data['schools_severe_risk_pct'] = round(risk_severe / total * 100, 1) if total > 0 else 0

    # Print national totals
    for scenario in ['census', 'state_trend']:
        for yr in [2030, 2035, 2040]:
            total_closures = sum(
                proj[scenario].get(yr, {}).get('expected_closures', 0)
                for proj in fertility_projections.values()
            )
            print(f"  {scenario} {yr}: ~{total_closures:.0f} expected closures nationally")

    return fertility_projections


# ============================================================
# Step 6: Download and simplify state GeoJSON
# ============================================================

def download_state_geojson():
    """Download Census state boundaries and convert to GeoJSON."""
    print("\n[Step 6] Processing state GeoJSON...")

    geojson_path = os.path.join(OUT_DIR, 'states.json')
    if os.path.exists(geojson_path):
        print(f"  states.json already exists ({os.path.getsize(geojson_path) / 1024:.0f} KB)")
        return True

    zip_path = os.path.join(DATA_DIR, 'cb_2022_us_state_20m.zip')
    url = 'https://www2.census.gov/geo/tiger/GENZ2022/shp/cb_2022_us_state_20m.zip'

    if not download_file(url, zip_path, "state boundaries"):
        return False

    # Try geopandas first
    try:
        import geopandas as gpd

        gdf = gpd.read_file(f"zip://{zip_path}")

        # Filter to 50 states + DC (exclude territories)
        valid_fips = set(STATE_FIPS.values())
        gdf = gdf[gdf['STATEFP'].isin(valid_fips)]

        # Simplify geometry
        gdf = gdf.to_crs('EPSG:4326')
        gdf['geometry'] = gdf['geometry'].simplify(tolerance=0.005, preserve_topology=True)

        # Keep only needed fields
        gdf = gdf[['STATEFP', 'STUSPS', 'NAME', 'geometry']]
        gdf.columns = ['FIPS', 'ABBR', 'NAME', 'geometry']

        gdf.to_file(geojson_path, driver='GeoJSON')
        print(f"  states.json: {len(gdf)} states, {os.path.getsize(geojson_path) / 1024:.0f} KB")
        return True

    except ImportError:
        print("  geopandas not available, trying manual approach...")

    # Manual fallback: extract shapefile from zip, use ogr2ogr if available
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(os.path.join(DATA_DIR, 'state_shp'))

        import subprocess
        shp_path = os.path.join(DATA_DIR, 'state_shp', 'cb_2022_us_state_20m.shp')
        result = subprocess.run(
            ['ogr2ogr', '-f', 'GeoJSON', '-simplify', '0.005', geojson_path, shp_path],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            # Filter to just states
            with open(geojson_path) as f:
                gj = json.load(f)
            valid_fips = set(STATE_FIPS.values())
            gj['features'] = [f for f in gj['features']
                              if f['properties'].get('STATEFP') in valid_fips]
            with open(geojson_path, 'w') as f:
                json.dump(gj, f)
            print(f"  states.json: {len(gj['features'])} states")
            return True
    except Exception as e:
        print(f"  Manual GeoJSON conversion failed: {e}")

    print("  ERROR: Could not create states.json. Install geopandas: pip install geopandas pyogrio")
    return False


# ============================================================
# Step 7: Output JSON files
# ============================================================

def output_json(fertility_projections, state_schools):
    """Write state_projections.json."""
    print("\n[Step 7] Writing output JSON...")

    # Build output structure
    output = {}

    for abbr, proj in fertility_projections.items():
        state_entry = {
            'state': proj['state'],
            'abbr': abbr,
            'fips': proj['fips'],
            'rate_2022': proj['rate_2022'],
            'decline_2005_2022': proj['decline_2005_2022'],
            'annual_decline_rate': proj['annual_decline_rate'],
            'current_schools': proj.get('current_schools', 0),
            'current_enrollment': proj.get('current_enrollment', 0),
            'buckets': proj.get('buckets', {}),
            'scenarios': {},
        }

        for scenario in ['census', 'state_trend']:
            scenario_output = {}
            for year in PROJECTION_YEARS:
                yr_data = proj[scenario].get(year, {})
                scenario_output[str(year)] = {
                    'fertility_rate': yr_data.get('rate', 0),
                    'birth_decline_pct': yr_data.get('birth_decline_pct', 0),
                    'post_2022_fraction': yr_data.get('post_2022_fraction', 0),
                    'enrollment_decline_pct': yr_data.get('enrollment_decline_pct', 0),
                    'projected_enrollment': yr_data.get('projected_enrollment', 0),
                    'expected_closures': yr_data.get('expected_closures', 0),
                    'expected_closures_pct': yr_data.get('expected_closures_pct', 0),
                    'schools_elevated_risk': yr_data.get('schools_elevated_risk', 0),
                    'schools_elevated_risk_pct': yr_data.get('schools_elevated_risk_pct', 0),
                    'schools_severe_risk': yr_data.get('schools_severe_risk', 0),
                    'schools_severe_risk_pct': yr_data.get('schools_severe_risk_pct', 0),
                }
            state_entry['scenarios'][scenario] = scenario_output

        output[abbr] = state_entry

    out_path = os.path.join(OUT_DIR, 'state_projections.json')
    with open(out_path, 'w') as f:
        json.dump(output, f)

    size_kb = os.path.getsize(out_path) / 1024
    print(f"  state_projections.json: {len(output)} states, {size_kb:.0f} KB")

    return output


# ============================================================
# Main
# ============================================================

def main():
    force = '--force' in sys.argv

    print("=" * 60)
    print("School Closure Risk Map — Data Processing")
    print("=" * 60)

    os.makedirs(DATA_DIR, exist_ok=True)

    # Step 1: CCD school data
    state_schools = process_ccd_schools()

    # Step 2: Census birth trajectory
    census_multipliers = extract_census_birth_trajectory()

    # Step 3: Fertility projections
    fertility_projections = build_fertility_projections(census_multipliers)

    # Step 4: Enrollment decline
    fertility_projections = compute_enrollment_decline(fertility_projections)

    # Step 5: Closure estimates
    fertility_projections = estimate_closures(fertility_projections, state_schools)

    # Step 6: State GeoJSON
    download_state_geojson()

    # Step 7: Output
    output_json(fertility_projections, state_schools)

    print("\n" + "=" * 60)
    print("Done! Output files:")
    print(f"  - state_projections.json")
    print(f"  - states.json")
    print("=" * 60)


if __name__ == '__main__':
    main()
