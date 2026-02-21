"""
Process Cooper Center National Population Projections (2020 baseline, 2030, 2040)
and add a 'cooper_center' scenario to state_projections.json.

Source: UVA Weldon Cooper Center for Public Service
URL: https://www.coopercenter.org/national-population-projections
File: NationalProjections_ProjectedAgeSexDistribution_2030-2050.xlsx
Released: July 1, 2024. Benchmarked to 2020 Decennial Census.

School-age population proxy = ages 5-9 + 10-14 + 3/5 × 15-19 (approximates K-12 ages 5-17).
Enrollment decline = change in school-age population from 2020 baseline, applied to NCES 2022 enrollment.
"""

import json
import os
import math
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# FIPS -> state abbreviation (standard 2-digit state FIPS)
FIPS_TO_ABBR = {
    '01': 'AL', '02': 'AK', '04': 'AZ', '05': 'AR', '06': 'CA',
    '08': 'CO', '09': 'CT', '10': 'DE', '11': 'DC', '12': 'FL',
    '13': 'GA', '15': 'HI', '16': 'ID', '17': 'IL', '18': 'IN',
    '19': 'IA', '20': 'KS', '21': 'KY', '22': 'LA', '23': 'ME',
    '24': 'MD', '25': 'MA', '26': 'MI', '27': 'MN', '28': 'MS',
    '29': 'MO', '30': 'MT', '31': 'NE', '32': 'NV', '33': 'NH',
    '34': 'NJ', '35': 'NM', '36': 'NY', '37': 'NC', '38': 'ND',
    '39': 'OH', '40': 'OK', '41': 'OR', '42': 'PA', '44': 'RI',
    '45': 'SC', '46': 'SD', '47': 'TN', '48': 'TX', '49': 'UT',
    '50': 'VT', '51': 'VA', '53': 'WA', '54': 'WV', '55': 'WI',
    '56': 'WY',
}

CLOSURE_PROBS = {
    'large_stable':   0.01,
    'medium_decline': 0.02,
    'small_steep':    0.05,
    'tiny':           0.10,
}


def extract_school_age(wb, sheet_name):
    """Extract ages 5-17 school-age population by state from one Cooper Center sheet."""
    ws = wb[sheet_name]
    state_data = {}

    for r in range(5, ws.max_row + 1):
        fips_raw = ws.cell(r, 1).value
        sex = ws.cell(r, 3).value

        if fips_raw is None or sex != 'Total' or fips_raw == 0:
            continue

        fips = str(int(round(fips_raw))).zfill(2)
        if fips not in FIPS_TO_ABBR:
            continue

        abbr = FIPS_TO_ABBR[fips]

        # col6 = 5-9, col7 = 10-14, col8 = 15-19
        pop_5_9   = float(ws.cell(r, 6).value or 0)
        pop_10_14 = float(ws.cell(r, 7).value or 0)
        pop_15_19 = float(ws.cell(r, 8).value or 0)

        # ages 5-17: full 5-9 and 10-14 bands, plus 3/5 of the 15-19 band
        school_age = pop_5_9 + pop_10_14 + (pop_15_19 * 3.0 / 5.0)
        state_data[abbr] = school_age

    return state_data


def compute_closures(decline_ratio, buckets, years_from_baseline):
    """Compute expected closures using same model as process_closure_risk.py."""
    avg_enrollment = {
        'under100': 55,
        '100_199':  150,
        '200_299':  250,
        '300_499':  400,
        '500plus':  750,
    }

    expected_closures = 0.0
    for bucket_name, count in buckets.items():
        if count == 0:
            continue

        avg_enr = avg_enrollment[bucket_name]
        projected_enr = avg_enr * (1.0 - decline_ratio)

        if projected_enr < 100:
            annual_prob = CLOSURE_PROBS['tiny']
        elif projected_enr < 200 and decline_ratio > 0.25:
            annual_prob = CLOSURE_PROBS['small_steep']
        elif projected_enr < 300 and decline_ratio > 0.10:
            annual_prob = CLOSURE_PROBS['medium_decline']
        elif decline_ratio > 0.10:
            annual_prob = CLOSURE_PROBS['medium_decline']
        else:
            annual_prob = CLOSURE_PROBS['large_stable']

        years = max(years_from_baseline, 1)
        cum_prob = 1.0 - (1.0 - annual_prob) ** years
        expected_closures += count * cum_prob

    return expected_closures


def main():
    xlsx_path = os.path.join(DATA_DIR, 'cooper_age_sex_2030_2050.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"ERROR: {xlsx_path} not found.")
        print("Download from: https://www.coopercenter.org/national-population-projections")
        return

    print("Loading Cooper Center Excel...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("Extracting school-age population (ages 5-17) by state and year...")
    school_age = {}
    for yr in [2020, 2030, 2040]:
        school_age[yr] = extract_school_age(wb, str(yr))
        total = sum(school_age[yr].values())
        print(f"  {yr}: {len(school_age[yr])} states/DC, national total = {total:,.0f}")

    # Load existing state_projections.json
    proj_path = os.path.join(OUT_DIR, 'state_projections.json')
    with open(proj_path, 'r') as f:
        state_proj = json.load(f)

    print(f"\nAdding Cooper Center scenario to {len(state_proj)} states...")
    added = 0
    skipped = 0

    for abbr, state_entry in state_proj.items():
        sa_2020 = school_age[2020].get(abbr)
        if not sa_2020 or sa_2020 == 0:
            print(f"  [{abbr}] No 2020 baseline, skipping")
            skipped += 1
            continue

        current_enrollment = state_entry.get('current_enrollment', 0)
        total_schools = state_entry.get('current_schools', 0)
        buckets = state_entry.get('buckets', {})

        scenario_data = {}

        # 2020 = Cooper Center baseline: 0% decline, no closures yet
        scenario_data['2020'] = {
            'enrollment_decline_pct': 0.0,
            'projected_enrollment': current_enrollment,
            'expected_closures': 0.0,
            'expected_closures_pct': 0.0,
            'cooper_school_age_2020': round(sa_2020),
            'cooper_school_age_proj': round(sa_2020),
        }

        for yr in [2030]:  # 2040 excluded: Cooper Center methodology unreliable at 20yr horizon
            sa_yr = school_age[yr].get(abbr)
            if not sa_yr:
                continue

            ratio = sa_yr / sa_2020
            decline_pct = round((1.0 - ratio) * 100, 2)

            proj_enrollment = round(current_enrollment * ratio)

            # Growing states expect net school openings, not closures — zero out.
            # Declining states: use actual decline ratio for closure model.
            if ratio >= 1.0:
                closures = 0.0
            else:
                closure_decline_ratio = 1.0 - ratio
                years_from_2025 = max(yr - 2025, 1)
                closures = compute_closures(closure_decline_ratio, buckets, years_from_2025)

            scenario_data[str(yr)] = {
                'enrollment_decline_pct': decline_pct,
                'projected_enrollment': proj_enrollment,
                'expected_closures': round(closures, 1),
                'expected_closures_pct': round(closures / total_schools * 100, 1) if total_schools > 0 else 0.0,
                'cooper_school_age_2020': round(sa_2020),
                'cooper_school_age_proj': round(sa_yr),
            }

        state_entry['scenarios']['cooper_center'] = scenario_data
        added += 1

    print(f"Added: {added}, skipped: {skipped}")

    print("\nSample results:")
    header = f"{'State':4}  {'2030 decline':>12}  {'2030 closures':>13}"
    print(header)
    print("-" * len(header))
    for abbr in ['MI', 'WI', 'TX', 'FL', 'WV', 'OH', 'CA', 'UT', 'MT', 'UT']:
        if abbr not in state_proj:
            continue
        s = state_proj[abbr]['scenarios'].get('cooper_center', {})
        d30 = s.get('2030', {})
        print(f"{abbr:4}  {d30.get('enrollment_decline_pct','?'):>11}%  "
              f"{d30.get('expected_closures','?'):>13}")

    # Write back
    with open(proj_path, 'w') as f:
        json.dump(state_proj, f)
    print(f"\nWrote updated state_projections.json ({os.path.getsize(proj_path)/1024:.0f} KB)")


if __name__ == '__main__':
    main()
