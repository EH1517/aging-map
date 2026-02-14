#!/usr/bin/env python3
"""
process_data.py — Data preparation for U.S. County-Level Aging Map

Usage:
    1. Download Hauer SSP data (SSP_asrc.csv) from https://dx.doi.org/10.17605/OSF.IO/9YNFC
       and place in data/
    2. Download Cooper Center projections from
       https://www.coopercenter.org/sites/default/files/2025-01/NationalProjections_ProjectedAgeSexDistribution_2030-2050.xlsx
       and place in data/
    3. Run: python process_data.py
    4. Open index.html in your browser
"""

import csv
import io
import json
import os
import sys
import zipfile
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

# Optional dependencies — give clear install instructions if missing
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR

HAUER_CSV = DATA_DIR / "SSP_asrc.csv"
HAUER_ZIP = DATA_DIR / "SSP_asrc.csv.zip"
HAUER_ZIP_ALT = DATA_DIR / "SSP_asrc.zip"
COOPER_XLSX = DATA_DIR / "NationalProjections_ProjectedAgeSexDistribution_2030-2050.xlsx"
COUNTY_ZIP = DATA_DIR / "cb_2022_us_county_500k.zip"
COUNTY_SHP_DIR = DATA_DIR / "cb_2022_us_county_500k"

YEARS = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
TARGET_YEARS = [2025, 2030, 2035, 2040, 2045, 2050]

# FIPS codes for states (including DC), excluding territories
STATE_FIPS = {
    "01", "02", "04", "05", "06", "08", "09", "10", "11", "12",
    "13", "15", "16", "17", "18", "19", "20", "21", "22", "23",
    "24", "25", "26", "27", "28", "29", "30", "31", "32", "33",
    "34", "35", "36", "37", "38", "39", "40", "41", "42", "44",
    "45", "46", "47", "48", "49", "50", "51", "53", "54", "55", "56"
}

# State FIPS to name mapping
STATE_FIPS_TO_ABBR = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY"
}

STATE_FIPS_TO_NAME = {
    "01": "Alabama", "02": "Alaska", "04": "Arizona", "05": "Arkansas",
    "06": "California", "08": "Colorado", "09": "Connecticut", "10": "Delaware",
    "11": "District of Columbia", "12": "Florida", "13": "Georgia", "15": "Hawaii",
    "16": "Idaho", "17": "Illinois", "18": "Indiana", "19": "Iowa",
    "20": "Kansas", "21": "Kentucky", "22": "Louisiana", "23": "Maine",
    "24": "Maryland", "25": "Massachusetts", "26": "Michigan", "27": "Minnesota",
    "28": "Mississippi", "29": "Missouri", "30": "Montana", "31": "Nebraska",
    "32": "Nevada", "33": "New Hampshire", "34": "New Jersey", "35": "New Mexico",
    "36": "New York", "37": "North Carolina", "38": "North Dakota", "39": "Ohio",
    "40": "Oklahoma", "41": "Oregon", "42": "Pennsylvania", "44": "Rhode Island",
    "45": "South Carolina", "46": "South Dakota", "47": "Tennessee", "48": "Texas",
    "49": "Utah", "50": "Vermont", "51": "Virginia", "53": "Washington",
    "54": "West Virginia", "55": "Wisconsin", "56": "Wyoming"
}


def check_files():
    """Check that required input files exist and give helpful messages."""
    missing = []
    if not HAUER_CSV.exists() and not HAUER_ZIP.exists() and not HAUER_ZIP_ALT.exists():
        missing.append(
            f"  - Hauer SSP data: Download SSP_asrc.csv.zip from\n"
            f"    https://dx.doi.org/10.17605/OSF.IO/9YNFC\n"
            f"    and place in {DATA_DIR}/\n"
            f"    (either the .csv or .csv.zip is fine)"
        )
    if not COOPER_XLSX.exists():
        missing.append(
            f"  - Cooper Center projections: Download from\n"
            f"    https://www.coopercenter.org/sites/default/files/2025-01/"
            f"NationalProjections_ProjectedAgeSexDistribution_2030-2050.xlsx\n"
            f"    and place in {DATA_DIR}/"
        )
    if missing:
        print("ERROR: Missing required data files:\n")
        print("\n".join(missing))
        print()
        sys.exit(1)


def download_county_boundaries():
    """Download Census county boundary shapefile if not present."""
    geojson_out = OUTPUT_DIR / "counties.json"
    if geojson_out.exists():
        print("  counties.json already exists, skipping download.")
        return

    url = "https://www2.census.gov/geo/tiger/GENZ2022/shp/cb_2022_us_county_500k.zip"
    if not COUNTY_ZIP.exists():
        print(f"  Downloading county boundaries from Census Bureau...")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "aging-map-builder/1.0"})
            with urllib.request.urlopen(req) as resp, open(COUNTY_ZIP, "wb") as f:
                f.write(resp.read())
            print(f"  Downloaded to {COUNTY_ZIP}")
        except urllib.error.URLError as e:
            print(f"  ERROR downloading county boundaries: {e}")
            print(f"  Please download manually from {url}")
            print(f"  and place in {DATA_DIR}/")
            sys.exit(1)


def convert_shapefile_to_geojson():
    """Convert the county shapefile to simplified GeoJSON."""
    geojson_out = OUTPUT_DIR / "counties.json"
    if geojson_out.exists():
        print("  counties.json already exists, skipping conversion.")
        return

    # Try geopandas first (works with pyogrio backend, no GDAL needed)
    try:
        import geopandas as gpd
        convert_with_geopandas(geojson_out)
        return
    except ImportError:
        pass

    # Fallback to fiona + shapely
    try:
        import fiona
        import shapely.geometry
    except ImportError:
        print("  ERROR: Need geopandas or fiona+shapely to convert shapefiles.")
        print("  Install with: pip install geopandas pyogrio shapely")
        print("  OR: Download a pre-made county GeoJSON and save as counties.json")
        sys.exit(1)

    print("  Converting shapefile to GeoJSON with fiona...")
    with zipfile.ZipFile(COUNTY_ZIP, "r") as z:
        z.extractall(COUNTY_SHP_DIR)

    shp_path = COUNTY_SHP_DIR / "cb_2022_us_county_500k.shp"
    features = []
    with fiona.open(str(shp_path)) as src:
        for feat in src:
            props = feat["properties"]
            state_fips = props.get("STATEFP", "")
            county_fips = props.get("COUNTYFP", "")
            if state_fips not in STATE_FIPS:
                continue
            fips = state_fips + county_fips
            geom = shapely.geometry.shape(feat["geometry"])
            geom = geom.simplify(0.005, preserve_topology=True)
            features.append({
                "type": "Feature",
                "properties": {
                    "GEOID": fips,
                    "NAME": props.get("NAME", ""),
                    "STATE": state_fips
                },
                "geometry": shapely.geometry.mapping(geom)
            })

    geojson = {"type": "FeatureCollection", "features": features}
    with open(geojson_out, "w") as f:
        json.dump(geojson, f)
    print(f"  Wrote {len(features)} county features to counties.json")


def convert_with_geopandas(geojson_out):
    """Fallback: convert using geopandas."""
    import geopandas as gpd

    print("  Converting shapefile to GeoJSON with geopandas...")
    with zipfile.ZipFile(COUNTY_ZIP, "r") as z:
        z.extractall(COUNTY_SHP_DIR)

    shp_path = COUNTY_SHP_DIR / "cb_2022_us_county_500k.shp"
    gdf = gpd.read_file(str(shp_path))

    # Filter to 50 states + DC
    gdf = gdf[gdf["STATEFP"].isin(STATE_FIPS)]

    # Simplify geometry
    gdf["geometry"] = gdf["geometry"].simplify(0.005, preserve_topology=True)

    # Keep only needed columns
    gdf = gdf.rename(columns={"GEOID": "GEOID", "NAME": "NAME", "STATEFP": "STATE"})
    gdf = gdf[["GEOID", "NAME", "STATE", "geometry"]]

    gdf.to_file(str(geojson_out), driver="GeoJSON")
    print(f"  Wrote {len(gdf)} county features to counties.json")


# ---------------------------------------------------------------------------
# Hauer data processing
# ---------------------------------------------------------------------------

def process_hauer():
    """Process Hauer SSP county-level projections into JSON."""
    print("\n[1/3] Processing Hauer county projections...")

    # Handle zip or csv
    csv_path = HAUER_CSV
    zip_path = HAUER_ZIP if HAUER_ZIP.exists() else (HAUER_ZIP_ALT if HAUER_ZIP_ALT.exists() else None)
    if not csv_path.exists() and zip_path:
        print(f"  Extracting {zip_path.name}...")
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(DATA_DIR)
        if not csv_path.exists():
            with zipfile.ZipFile(zip_path, "r") as z:
                names = z.namelist()
                csv_files = [n for n in names if n.endswith(".csv")]
                if csv_files:
                    extracted = DATA_DIR / csv_files[0]
                    if extracted.exists():
                        csv_path = extracted

    if not csv_path.exists():
        print(f"  ERROR: Cannot find SSP_asrc.csv in {DATA_DIR}")
        sys.exit(1)

    # Parse CSV: aggregate population by county, year, and age group
    # Actual structure: YEAR, SEX, STATE, COUNTY, GEOID, RACE, AGE, SSP1, SSP2, SSP3, SSP4, SSP5
    # AGE is 1-18 representing 5-year groups: 1=0-4, 2=5-9, ..., 14=65-69, 15=70-74, 16=75-79, 17=80-84, 18=85+
    # So AGE >= 14 means 65+, AGE == 18 means 85+
    # SSP2 is a column containing the population value for that scenario
    print("  Reading Hauer CSV (this may take a moment for a large file)...")

    # county_data[fips][year] = {"total": 0, "65plus": 0, "85plus": 0}
    county_data = defaultdict(lambda: defaultdict(lambda: {"total": 0.0, "65plus": 0.0, "85plus": 0.0}))

    row_count = 0
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [fn.strip() for fn in reader.fieldnames]

        for row in reader:
            row_count += 1
            if row_count % 5000000 == 0:
                print(f"    ...processed {row_count:,} rows")

            year = int(row.get("YEAR", "0"))
            if year not in YEARS and year != 2020:
                continue

            state_fips = row.get("STATE", "").strip().zfill(2)
            county_fips = row.get("COUNTY", "").strip().zfill(3)
            if state_fips not in STATE_FIPS:
                continue

            fips = state_fips + county_fips
            age_group = int(row.get("AGE", "0"))
            value = float(row.get("SSP2", "0"))

            county_data[fips][year]["total"] += value
            if age_group >= 14:  # 14=65-69, 15=70-74, 16=75-79, 17=80-84, 18=85+
                county_data[fips][year]["65plus"] += value
            if age_group >= 18:  # 18=85+
                county_data[fips][year]["85plus"] += value

    print(f"  Processed {row_count:,} total rows, found {len(county_data)} counties")

    # Build output JSON
    # Interpolate for years not directly in the data
    hauer_years_available = sorted({y for fips in county_data for y in county_data[fips]})
    print(f"  Available years in Hauer data: {hauer_years_available}")

    output = {}
    for fips, years_data in county_data.items():
        state_abbr = STATE_FIPS_TO_ABBR.get(fips[:2], fips[:2])
        entry = {
            "name": f"{fips}, {state_abbr}",  # placeholder, will be enriched
            "years": {}
        }

        for target_year in TARGET_YEARS:
            if target_year in years_data:
                d = years_data[target_year]
            else:
                # Linear interpolation
                avail = sorted(years_data.keys())
                before = [y for y in avail if y <= target_year]
                after = [y for y in avail if y >= target_year]
                if before and after:
                    y1, y2 = before[-1], after[0]
                    if y1 == y2:
                        d = years_data[y1]
                    else:
                        t = (target_year - y1) / (y2 - y1)
                        d1, d2 = years_data[y1], years_data[y2]
                        d = {
                            "total": d1["total"] + t * (d2["total"] - d1["total"]),
                            "65plus": d1["65plus"] + t * (d2["65plus"] - d1["65plus"]),
                            "85plus": d1["85plus"] + t * (d2["85plus"] - d1["85plus"]),
                        }
                elif before:
                    d = years_data[before[-1]]
                elif after:
                    d = years_data[after[0]]
                else:
                    continue

            total = d["total"]
            p65 = d["65plus"]
            p85 = d["85plus"]
            if total > 0:
                entry["years"][str(target_year)] = {
                    "pct_65plus": round(100 * p65 / total, 1),
                    "pct_85plus": round(100 * p85 / total, 1),
                    "pop_total": round(total),
                    "pop_65plus": round(p65),
                    "pop_85plus": round(p85),
                }

        if entry["years"]:
            output[fips] = entry

    out_path = OUTPUT_DIR / "hauer_data.json"
    with open(out_path, "w") as f:
        json.dump(output, f)
    print(f"  Wrote {len(output)} counties to hauer_data.json")
    return output


# ---------------------------------------------------------------------------
# Cooper Center data processing
# ---------------------------------------------------------------------------

def fetch_acs_data():
    """Fetch ACS county-level age data from Census API or use cached file."""
    acs_cache = DATA_DIR / "acs_county_age.json"
    if acs_cache.exists():
        print("  Using cached ACS data from data/acs_county_age.json")
        with open(acs_cache) as f:
            return json.load(f)

    print("  Fetching ACS county-level age data from Census API...")
    print("  (This requires internet access. If it fails, see instructions below.)")

    # B01001: Sex by Age
    # We need: total pop, 65-66, 67-69, 70-74, 75-79, 80-84, 85+
    # Male 65+: B01001_020E through B01001_025E
    # Female 65+: B01001_044E through B01001_049E
    # Male 85+: B01001_025E
    # Female 85+: B01001_049E
    male_65plus = ["B01001_020E", "B01001_021E", "B01001_022E", "B01001_023E", "B01001_024E", "B01001_025E"]
    female_65plus = ["B01001_044E", "B01001_045E", "B01001_046E", "B01001_047E", "B01001_048E", "B01001_049E"]
    male_85plus = ["B01001_025E"]
    female_85plus = ["B01001_049E"]

    all_vars = ["NAME", "B01001_001E"] + male_65plus + female_65plus
    var_str = ",".join(all_vars)

    url = f"https://api.census.gov/data/2022/acs/acs5?get={var_str}&for=county:*&in=state:*"

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "aging-map-builder/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode())
    except Exception as e:
        print(f"  ERROR fetching ACS data: {e}")
        print(f"  You can manually download the data by visiting:")
        print(f"  {url}")
        print(f"  Save the result as {acs_cache}")
        sys.exit(1)

    headers = data[0]
    rows = data[1:]

    # Parse into county records
    county_age = {}
    for row in rows:
        record = dict(zip(headers, row))
        state = record["state"]
        county = record["county"]
        if state not in STATE_FIPS:
            continue

        fips = state + county
        total_pop = int(record.get("B01001_001E", 0) or 0)

        pop_65plus = 0
        for v in male_65plus + female_65plus:
            pop_65plus += int(record.get(v, 0) or 0)

        pop_85plus = 0
        for v in male_85plus + female_85plus:
            pop_85plus += int(record.get(v, 0) or 0)

        name = record.get("NAME", fips)

        county_age[fips] = {
            "name": name,
            "state_fips": state,
            "pop_total": total_pop,
            "pop_65plus": pop_65plus,
            "pop_85plus": pop_85plus,
        }

    # Cache it
    with open(acs_cache, "w") as f:
        json.dump(county_age, f)
    print(f"  Fetched and cached ACS data for {len(county_age)} counties")
    return county_age


def process_cooper():
    """Process Cooper Center state projections distributed to counties."""
    print("\n[2/3] Processing Cooper Center projections...")

    # Read Cooper Center Excel
    # Structure per sheet (2020, 2030, 2040, 2050):
    #   Row 0: Title
    #   Row 1: Source
    #   Row 2: Headers — FIPS, Geography Name, Sex, Total Population, "Population by Age, YYYY (Number)"
    #   Row 3: Age sub-headers — (blanks), then "0 to 4", "5 to 9", ..., "80 to 84", "85+"
    #   Row 4+: Data rows. FIPS=integer, Sex="Total"/"Male"/"Female"
    print("  Reading Cooper Center Excel file...")
    wb = openpyxl.load_workbook(str(COOPER_XLSX), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Available sheets: {sheet_names}")

    state_projections = {}  # state_fips -> {year -> {"total": X, "65plus": Y, "85plus": Z}}

    for sheet_name in sheet_names:
        # Extract year from sheet name
        year = None
        for y in [2020, 2030, 2040, 2050]:
            if str(y) == sheet_name.strip():
                year = y
                break
        if year is None:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Row 3 (index 3) has age group sub-headers
        age_row = rows[3]
        age_col_map = {}  # col_index -> age_lower_bound
        for j, cell in enumerate(age_row):
            if cell is None:
                continue
            age = parse_age_group(str(cell).strip())
            if age is not None:
                age_col_map[j] = age

        print(f"  Sheet '{sheet_name}': {len(age_col_map)} age columns, ages {sorted(age_col_map.values())}")

        # Data rows start at index 4
        # Columns: 0=FIPS, 1=Geography Name, 2=Sex, 3=Total Population, 4+=age groups
        for row in rows[4:]:
            if not row or row[0] is None:
                continue

            # Only "Total" sex rows
            sex = str(row[2]).strip() if row[2] else ""
            if sex.lower() != "total":
                continue

            state_name = str(row[1]).strip() if row[1] else ""
            fips_val = row[0]

            # Skip US total (FIPS=0)
            try:
                fips_int = int(fips_val)
            except (ValueError, TypeError):
                continue
            if fips_int == 0:
                continue

            # Get state FIPS from the FIPS column
            state_fips = str(fips_int).zfill(2)
            if state_fips not in STATE_FIPS:
                continue

            total_pop = float(row[3]) if row[3] else 0
            pop_65plus = 0
            pop_85plus = 0
            for col_idx, age_lower in age_col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        val = float(row[col_idx])
                    except (ValueError, TypeError):
                        continue
                    if age_lower >= 65:
                        pop_65plus += val
                    if age_lower >= 85:
                        pop_85plus += val

            if total_pop <= 0:
                continue

            if state_fips not in state_projections:
                state_projections[state_fips] = {}
            state_projections[state_fips][year] = {
                "total": total_pop,
                "65plus": pop_65plus,
                "85plus": pop_85plus,
            }

    wb.close()
    print(f"  Parsed state projections for {len(state_projections)} states")
    for sf, yrs in sorted(state_projections.items()):
        st = STATE_FIPS_TO_ABBR.get(sf, sf)
        yr_keys = sorted(yrs.keys())
        if yr_keys:
            sample = yrs[yr_keys[0]]
            print(f"    {st}: years={yr_keys}, sample total={sample['total']:,.0f}, 65+={sample['65plus']:,.0f}")

    # Fetch ACS county data
    county_age = fetch_acs_data()

    # Compute county shares of state 65+ and 85+ populations
    state_totals = defaultdict(lambda: {"pop_total": 0, "pop_65plus": 0, "pop_85plus": 0})
    for fips, cdata in county_age.items():
        sf = cdata["state_fips"]
        state_totals[sf]["pop_total"] += cdata["pop_total"]
        state_totals[sf]["pop_65plus"] += cdata["pop_65plus"]
        state_totals[sf]["pop_85plus"] += cdata["pop_85plus"]

    # Distribute state projections to counties
    output = {}
    for fips, cdata in county_age.items():
        sf = cdata["state_fips"]
        st = state_totals[sf]

        if sf not in state_projections:
            continue

        # County shares (from ACS baseline)
        share_total = cdata["pop_total"] / st["pop_total"] if st["pop_total"] > 0 else 0
        share_65 = cdata["pop_65plus"] / st["pop_65plus"] if st["pop_65plus"] > 0 else 0
        share_85 = cdata["pop_85plus"] / st["pop_85plus"] if st["pop_85plus"] > 0 else 0

        # ACS baseline (~2020)
        baseline = {
            "total": cdata["pop_total"],
            "65plus": cdata["pop_65plus"],
            "85plus": cdata["pop_85plus"],
        }

        # Get projected values for each Cooper Center year
        proj_years = {}  # year -> {total, 65plus, 85plus}
        proj_years[2020] = baseline  # Use ACS as 2020 baseline

        for yr, sp in state_projections[sf].items():
            proj_years[yr] = {
                "total": sp["total"] * share_total,
                "65plus": sp["65plus"] * share_65,
                "85plus": sp["85plus"] * share_85,
            }

        # Interpolate for all target years
        avail_years = sorted(proj_years.keys())
        state_abbr = STATE_FIPS_TO_ABBR.get(sf, sf)
        entry = {
            "name": cdata["name"],
            "years": {}
        }

        for target_year in TARGET_YEARS:
            if target_year in proj_years:
                d = proj_years[target_year]
            else:
                before = [y for y in avail_years if y <= target_year]
                after = [y for y in avail_years if y >= target_year]
                if before and after:
                    y1, y2 = before[-1], after[0]
                    if y1 == y2:
                        d = proj_years[y1]
                    else:
                        t = (target_year - y1) / (y2 - y1)
                        d1, d2 = proj_years[y1], proj_years[y2]
                        d = {
                            "total": d1["total"] + t * (d2["total"] - d1["total"]),
                            "65plus": d1["65plus"] + t * (d2["65plus"] - d1["65plus"]),
                            "85plus": d1["85plus"] + t * (d2["85plus"] - d1["85plus"]),
                        }
                elif before:
                    d = proj_years[before[-1]]
                elif after:
                    d = proj_years[after[0]]
                else:
                    continue

            total = d["total"]
            p65 = d["65plus"]
            p85 = d["85plus"]
            if total > 0:
                entry["years"][str(target_year)] = {
                    "pct_65plus": round(100 * p65 / total, 1),
                    "pct_85plus": round(100 * p85 / total, 1),
                    "pop_total": round(total),
                    "pop_65plus": round(p65),
                    "pop_85plus": round(p85),
                }

        if entry["years"]:
            output[fips] = entry

    out_path = OUTPUT_DIR / "cooper_data.json"
    with open(out_path, "w") as f:
        json.dump(output, f)
    print(f"  Wrote {len(output)} counties to cooper_data.json")
    return output


def parse_age_group(label):
    """Parse age group label to lower bound. Returns None if not an age group."""
    if not label:
        return None
    label = label.strip()

    # Common patterns: "0-4", "5-9", "65-69", "85+", "85 and over",
    # "Under 5", "0 to 4", "85 and Over", "85+"
    import re

    # "85+", "85 and over", "85 and older", "85 plus"
    m = re.match(r"^(\d+)\s*[\+]", label)
    if m:
        return int(m.group(1))

    m = re.match(r"^(\d+)\s*(?:and\s*(?:over|older|above)|plus)", label, re.IGNORECASE)
    if m:
        return int(m.group(1))

    # "0-4", "65-69", "0 - 4"
    m = re.match(r"^(\d+)\s*[-–—]\s*(\d+)", label)
    if m:
        return int(m.group(1))

    # "0 to 4", "65 to 69"
    m = re.match(r"^(\d+)\s*to\s*(\d+)", label, re.IGNORECASE)
    if m:
        return int(m.group(1))

    # "Under 5"
    m = re.match(r"^under\s*(\d+)", label, re.IGNORECASE)
    if m:
        return 0

    # Just a number (some datasets use single age)
    m = re.match(r"^(\d+)$", label)
    if m:
        val = int(m.group(1))
        if 0 <= val <= 100:
            return val

    return None


def state_name_to_fips(name):
    """Convert state name or abbreviation to FIPS code."""
    name = name.strip()
    # Check abbreviation
    for fips, abbr in STATE_FIPS_TO_ABBR.items():
        if name.upper() == abbr:
            return fips
    # Check full name
    for fips, full_name in STATE_FIPS_TO_NAME.items():
        if name.lower() == full_name.lower():
            return fips
    # Partial match
    for fips, full_name in STATE_FIPS_TO_NAME.items():
        if name.lower() in full_name.lower() or full_name.lower() in name.lower():
            return fips
    return None


# ---------------------------------------------------------------------------
# County GeoJSON processing
# ---------------------------------------------------------------------------

def process_geojson():
    """Download and process county boundary GeoJSON."""
    print("\n[3/3] Processing county boundaries...")
    download_county_boundaries()
    convert_shapefile_to_geojson()

    geojson_path = OUTPUT_DIR / "counties.json"
    if not geojson_path.exists():
        print("  ERROR: counties.json not found. Please provide a county GeoJSON file.")
        print("  You can download one from: https://eric.clst.org/tech/usgeojson/")
        print("  Save it as counties.json in the project root.")
        sys.exit(1)


def enrich_names(hauer_data, cooper_data):
    """Use Cooper Center names (from ACS) to enrich Hauer data county names."""
    for fips, entry in hauer_data.items():
        if fips in cooper_data and cooper_data[fips].get("name"):
            entry["name"] = cooper_data[fips]["name"]

    # Re-save hauer data
    out_path = OUTPUT_DIR / "hauer_data.json"
    with open(out_path, "w") as f:
        json.dump(hauer_data, f)
    print(f"\n  Enriched Hauer county names from ACS data")


def main():
    print("=" * 60)
    print("  U.S. County-Level Aging Map — Data Processor")
    print("=" * 60)

    check_files()

    hauer_data = process_hauer()
    cooper_data = process_cooper()
    enrich_names(hauer_data, cooper_data)
    process_geojson()

    print("\n" + "=" * 60)
    print("  DONE! All data files generated.")
    print("=" * 60)
    print()
    print("  Output files:")
    print(f"    - hauer_data.json   ({(OUTPUT_DIR / 'hauer_data.json').stat().st_size / 1024:.0f} KB)")
    print(f"    - cooper_data.json  ({(OUTPUT_DIR / 'cooper_data.json').stat().st_size / 1024:.0f} KB)")
    geojson_path = OUTPUT_DIR / "counties.json"
    if geojson_path.exists():
        print(f"    - counties.json     ({geojson_path.stat().st_size / 1024:.0f} KB)")
    print()
    print("  Next step: Open index.html in your browser!")
    print()


if __name__ == "__main__":
    main()
